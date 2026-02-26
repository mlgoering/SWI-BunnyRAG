from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path
from statistics import mean, median, pstdev
from typing import Dict, List


def _load_rows(path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            parsed: Dict[str, object] = {}
            for k, v in row.items():
                if v in ("N/A", "", None):
                    parsed[k] = "N/A"
                    continue
                if k.startswith("pass_"):
                    parsed[k] = str(v).strip().lower() == "true"
                    continue
                if k in {
                    "dataset_seed",
                    "query_seed",
                    "top_k",
                    "seed_k",
                    "communities_total",
                    "selected_k",
                    "represented_count",
                }:
                    parsed[k] = int(float(v))
                    continue
                if k in {"dataset_id", "query_id", "method"}:
                    parsed[k] = v
                    continue
                if k == "lambda":
                    parsed[k] = v if v == "NA" else float(v)
                    continue
                parsed[k] = float(v)
            rows.append(parsed)
    return rows


def _sheet_title(name: str) -> str:
    return name[:31]


def _to_float_or_none(value: object) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    return None


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))


def _is_bunny(row: Dict[str, object]) -> bool:
    return row.get("method") == "bunny"


def _bool_status(value: object):
    if isinstance(value, bool):
        return value
    return None


def _metric_pass_statuses(row: Dict[str, object], thresholds: Dict[str, float]) -> Dict[str, bool | None]:
    if not _is_bunny(row):
        return {
            "lift_coverage_vs_graphrag": None,
            "lift_entropy_vs_graphrag": None,
            "delta_entropy_vs_graphrag": None,
            "improve_max_share_vs_graphrag": None,
            "improve_kl_vs_graphrag": None,
            "sim_retention_vs_graphrag": None,
            "lift_coverage_vs_random": None,
            "lift_entropy_vs_random": None,
            "delta_entropy_vs_random": None,
            "improve_max_share_vs_random": None,
            "improve_kl_vs_random": None,
            "sim_retention_vs_random": None,
        }
    lift_entropy_gr = _to_float_or_none(row["lift_entropy_vs_graphrag"])
    lift_entropy_rd = _to_float_or_none(row["lift_entropy_vs_random"])
    return {
        "lift_coverage_vs_graphrag": float(row["lift_coverage_vs_graphrag"])
        >= thresholds["min_lift_coverage_vs_graphrag"],
        "lift_entropy_vs_graphrag": (
            None
            if lift_entropy_gr is None
            else lift_entropy_gr >= thresholds["min_lift_entropy_vs_graphrag"]
        ),
        "delta_entropy_vs_graphrag": float(row["delta_entropy_vs_graphrag"]) >= 0.0,
        "improve_max_share_vs_graphrag": float(row["improve_max_share_vs_graphrag"])
        >= thresholds["min_improve_max_share_vs_graphrag"],
        "improve_kl_vs_graphrag": float(row["improve_kl_vs_graphrag"])
        >= thresholds["min_improve_kl_vs_graphrag"],
        "sim_retention_vs_graphrag": float(row["sim_retention_vs_graphrag"])
        >= thresholds["min_sim_retention_vs_graphrag"],
        "lift_coverage_vs_random": float(row["lift_coverage_vs_random"])
        >= thresholds["min_lift_coverage_vs_random"],
        "lift_entropy_vs_random": (
            None
            if lift_entropy_rd is None
            else lift_entropy_rd >= thresholds["min_lift_entropy_vs_random"]
        ),
        "delta_entropy_vs_random": float(row["delta_entropy_vs_random"]) >= 0.0,
        "improve_max_share_vs_random": float(row["improve_max_share_vs_random"])
        >= thresholds["min_improve_max_share_vs_random"],
        "improve_kl_vs_random": float(row["improve_kl_vs_random"])
        >= thresholds["min_improve_kl_vs_random"],
        "sim_retention_vs_random": float(row["sim_retention_vs_random"])
        >= thresholds["min_sim_retention_vs_random"],
    }


def _write_table(ws, start_row: int, headers: List[str], rows: List[Dict[str, object]]) -> int:
    for c, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=c, value=h)
    r = start_row + 1
    for row in rows:
        for c, h in enumerate(headers, start=1):
            ws.cell(row=r, column=c, value=row.get(h))
        r += 1
    return r


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build behavior-test Excel workbook from runner CSV outputs."
    )
    parser.add_argument(
        "--rows-csv",
        default="synthetic_bunny/output/behavior_runner/behavior_results_rows.csv",
        help="Path to behavior_results_rows.csv.",
    )
    parser.add_argument(
        "--metadata-json",
        default="synthetic_bunny/output/behavior_runner/behavior_metadata.json",
        help="Path to behavior_metadata.json.",
    )
    parser.add_argument(
        "--output-xlsx",
        default="presentation/testing/behavior_report.xlsx",
        help="Output workbook path.",
    )
    args = parser.parse_args()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    rows_path = Path(args.rows_csv)
    metadata_path = Path(args.metadata_json)
    rows = _load_rows(rows_path)
    meta = json.loads(metadata_path.read_text(encoding="utf-8"))
    thresholds: Dict[str, float] = meta.get("thresholds", {})
    lambdas = [float(x) for x in meta.get("lambdas", [])]

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "00_cumulative"

    # Use opaque ARGB colors so Excel renders fills visibly.
    green = PatternFill(fill_type="solid", fgColor="FFC6EFCE")
    red = PatternFill(fill_type="solid", fgColor="FFFFC7CE")
    gray = PatternFill(fill_type="solid", fgColor="FFEFEFEF")

    ws0["A1"] = "Thresholds"
    threshold_items = list(thresholds.items())
    for i, (k, v) in enumerate(threshold_items, start=2):
        ws0.cell(row=i, column=1, value=k)
        ws0.cell(row=i, column=2, value=v)

    params = meta.get("params", {})
    ws0["D1"] = "Run Setup"
    ws0["D2"] = "n"
    ws0["E2"] = params.get("n")
    ws0["D3"] = "dim"
    ws0["E3"] = params.get("dim")
    ws0["D4"] = "scale_prob"
    ws0["E4"] = params.get("scale_prob")

    legend_row = len(threshold_items) + 3
    ws0.cell(row=legend_row, column=1, value="Legend")
    ws0.cell(row=legend_row + 1, column=1, value="Pass")
    ws0.cell(row=legend_row + 2, column=1, value="Fail")
    ws0.cell(row=legend_row + 1, column=2, value="Green")
    ws0.cell(row=legend_row + 2, column=2, value="Red")
    ws0.cell(row=legend_row + 1, column=2).fill = green
    ws0.cell(row=legend_row + 2, column=2).fill = red

    bunny_rows = [r for r in rows if r.get("method") == "bunny"]
    by_lambda: Dict[float, List[Dict[str, object]]] = defaultdict(list)
    for row in bunny_rows:
        by_lambda[float(row["lambda"])].append(row)

    summary_headers = [
        "lambda",
        "runs",
        "mean_lift_entropy_vs_graphrag",
        "mean_lift_entropy_vs_random",
        "mean_delta_entropy_vs_graphrag",
        "mean_delta_entropy_vs_random",
        "median_lift_entropy_vs_graphrag",
        "median_lift_entropy_vs_random",
        "std_lift_entropy_vs_graphrag",
        "std_lift_entropy_vs_random",
        "mean_lift_coverage_vs_graphrag",
        "mean_lift_coverage_vs_random",
        "mean_improve_kl_vs_graphrag",
        "mean_improve_kl_vs_random",
        "mean_improve_max_share_vs_graphrag",
        "mean_improve_max_share_vs_random",
        "mean_sim_retention_vs_graphrag",
        "mean_sim_retention_vs_random",
        "pass_mean_diversity_vs_graphrag",
        "pass_mean_diversity_vs_random",
        "pass_mean_relevance_vs_graphrag",
        "pass_mean_relevance_vs_random",
        "pass_mean_overall_vs_graphrag",
        "pass_mean_overall_vs_random",
        "pass_rate_overall_vs_graphrag",
        "pass_rate_overall_vs_random",
    ]
    summary_rows: List[Dict[str, object]] = []
    for lam in sorted(lambdas):
        lam_rows = by_lambda.get(lam, [])
        if not lam_rows:
            continue
        lift_entropy_vals = [
            float(r["lift_entropy_vs_graphrag"])
            for r in lam_rows
            if isinstance(r["lift_entropy_vs_graphrag"], (float, int))
        ]
        lift_entropy_vals_rd = [
            float(r["lift_entropy_vs_random"])
            for r in lam_rows
            if isinstance(r["lift_entropy_vs_random"], (float, int))
        ]
        mean_lift_ent = "N/A" if not lift_entropy_vals else mean(lift_entropy_vals)
        mean_lift_ent_rd = (
            "N/A" if not lift_entropy_vals_rd else mean(lift_entropy_vals_rd)
        )
        median_lift_ent = "N/A" if not lift_entropy_vals else median(lift_entropy_vals)
        median_lift_ent_rd = (
            "N/A" if not lift_entropy_vals_rd else median(lift_entropy_vals_rd)
        )
        std_lift_ent = (
            "N/A"
            if not lift_entropy_vals
            else pstdev(lift_entropy_vals)
            if len(lift_entropy_vals) > 1
            else 0.0
        )
        std_lift_ent_rd = (
            "N/A"
            if not lift_entropy_vals_rd
            else pstdev(lift_entropy_vals_rd)
            if len(lift_entropy_vals_rd) > 1
            else 0.0
        )
        pass_gr_vals = [bool(r["pass_overall_vs_graphrag"]) for r in lam_rows]
        pass_rd_vals = [bool(r["pass_overall_vs_random"]) for r in lam_rows]
        mean_lift_cov = mean(float(r["lift_coverage_vs_graphrag"]) for r in lam_rows)
        mean_lift_cov_rd = mean(float(r["lift_coverage_vs_random"]) for r in lam_rows)
        mean_delta_ent = mean(float(r["delta_entropy_vs_graphrag"]) for r in lam_rows)
        mean_delta_ent_rd = mean(float(r["delta_entropy_vs_random"]) for r in lam_rows)
        mean_imp_kl = mean(float(r["improve_kl_vs_graphrag"]) for r in lam_rows)
        mean_imp_kl_rd = mean(float(r["improve_kl_vs_random"]) for r in lam_rows)
        mean_imp_share = mean(
            float(r["improve_max_share_vs_graphrag"]) for r in lam_rows
        )
        mean_imp_share_rd = mean(
            float(r["improve_max_share_vs_random"]) for r in lam_rows
        )
        mean_sim_ret = mean(float(r["sim_retention_vs_graphrag"]) for r in lam_rows)
        mean_sim_ret_rd = mean(float(r["sim_retention_vs_random"]) for r in lam_rows)
        pass_div = (
            mean_lift_cov >= thresholds["min_lift_coverage_vs_graphrag"]
            and mean_delta_ent >= 0.0
            and mean_imp_kl >= thresholds["min_improve_kl_vs_graphrag"]
            and mean_imp_share >= thresholds["min_improve_max_share_vs_graphrag"]
        )
        pass_rel = mean_sim_ret >= thresholds["min_sim_retention_vs_graphrag"]
        pass_div_rd = (
            mean_lift_cov_rd >= thresholds["min_lift_coverage_vs_random"]
            and mean_delta_ent_rd >= 0.0
            and mean_imp_kl_rd >= thresholds["min_improve_kl_vs_random"]
            and mean_imp_share_rd >= thresholds["min_improve_max_share_vs_random"]
        )
        pass_rel_rd = mean_sim_ret_rd >= thresholds["min_sim_retention_vs_random"]
        summary_rows.append(
            {
                "lambda": lam,
                "runs": len(lam_rows),
                "mean_lift_entropy_vs_graphrag": mean_lift_ent,
                "mean_lift_entropy_vs_random": mean_lift_ent_rd,
                "mean_delta_entropy_vs_graphrag": mean_delta_ent,
                "mean_delta_entropy_vs_random": mean_delta_ent_rd,
                "median_lift_entropy_vs_graphrag": median_lift_ent,
                "median_lift_entropy_vs_random": median_lift_ent_rd,
                "std_lift_entropy_vs_graphrag": std_lift_ent,
                "std_lift_entropy_vs_random": std_lift_ent_rd,
                "mean_lift_coverage_vs_graphrag": mean_lift_cov,
                "mean_lift_coverage_vs_random": mean_lift_cov_rd,
                "mean_improve_kl_vs_graphrag": mean_imp_kl,
                "mean_improve_kl_vs_random": mean_imp_kl_rd,
                "mean_improve_max_share_vs_graphrag": mean_imp_share,
                "mean_improve_max_share_vs_random": mean_imp_share_rd,
                "mean_sim_retention_vs_graphrag": mean_sim_ret,
                "mean_sim_retention_vs_random": mean_sim_ret_rd,
                "pass_mean_diversity_vs_graphrag": pass_div,
                "pass_mean_diversity_vs_random": pass_div_rd,
                "pass_mean_relevance_vs_graphrag": pass_rel,
                "pass_mean_relevance_vs_random": pass_rel_rd,
                "pass_mean_overall_vs_graphrag": bool(pass_div and pass_rel),
                "pass_mean_overall_vs_random": bool(pass_div_rd and pass_rel_rd),
                "pass_rate_overall_vs_graphrag": sum(pass_gr_vals) / len(pass_gr_vals),
                "pass_rate_overall_vs_random": sum(pass_rd_vals) / len(pass_rd_vals),
            }
        )

    summary_start = legend_row + 5
    ws0.cell(row=summary_start - 1, column=1, value="Cumulative by Lambda")
    next_row = _write_table(ws0, summary_start, summary_headers, summary_rows)
    summary_col = {h: i + 1 for i, h in enumerate(summary_headers)}
    for ridx, row in enumerate(summary_rows, start=summary_start + 1):
        metric_checks = {
            "mean_lift_entropy_vs_graphrag": (
                None
                if not isinstance(row["mean_lift_entropy_vs_graphrag"], (float, int))
                else row["mean_lift_entropy_vs_graphrag"]
                >= thresholds["min_lift_entropy_vs_graphrag"]
            ),
            "mean_lift_entropy_vs_random": (
                None
                if not isinstance(row["mean_lift_entropy_vs_random"], (float, int))
                else row["mean_lift_entropy_vs_random"]
                >= thresholds["min_lift_entropy_vs_random"]
            ),
            "mean_delta_entropy_vs_graphrag": row["mean_delta_entropy_vs_graphrag"] >= 0.0,
            "mean_delta_entropy_vs_random": row["mean_delta_entropy_vs_random"] >= 0.0,
            "mean_lift_coverage_vs_graphrag": row["mean_lift_coverage_vs_graphrag"]
            >= thresholds["min_lift_coverage_vs_graphrag"],
            "mean_lift_coverage_vs_random": row["mean_lift_coverage_vs_random"]
            >= thresholds["min_lift_coverage_vs_random"],
            "mean_improve_kl_vs_graphrag": row["mean_improve_kl_vs_graphrag"]
            >= thresholds["min_improve_kl_vs_graphrag"],
            "mean_improve_kl_vs_random": row["mean_improve_kl_vs_random"]
            >= thresholds["min_improve_kl_vs_random"],
            "mean_improve_max_share_vs_graphrag": row[
                "mean_improve_max_share_vs_graphrag"
            ]
            >= thresholds["min_improve_max_share_vs_graphrag"],
            "mean_improve_max_share_vs_random": row[
                "mean_improve_max_share_vs_random"
            ]
            >= thresholds["min_improve_max_share_vs_random"],
            "mean_sim_retention_vs_graphrag": row["mean_sim_retention_vs_graphrag"]
            >= thresholds["min_sim_retention_vs_graphrag"],
            "mean_sim_retention_vs_random": row["mean_sim_retention_vs_random"]
            >= thresholds["min_sim_retention_vs_random"],
            "pass_mean_diversity_vs_graphrag": bool(
                row["pass_mean_diversity_vs_graphrag"]
            ),
            "pass_mean_diversity_vs_random": bool(
                row["pass_mean_diversity_vs_random"]
            ),
            "pass_mean_relevance_vs_graphrag": bool(
                row["pass_mean_relevance_vs_graphrag"]
            ),
            "pass_mean_relevance_vs_random": bool(
                row["pass_mean_relevance_vs_random"]
            ),
            "pass_mean_overall_vs_graphrag": bool(
                row["pass_mean_overall_vs_graphrag"]
            ),
            "pass_mean_overall_vs_random": bool(row["pass_mean_overall_vs_random"]),
        }
        for col_name, ok in metric_checks.items():
            if ok is None:
                continue
            ws0.cell(row=ridx, column=summary_col[col_name]).fill = green if ok else red

    for c, h in enumerate(summary_headers, start=1):
        if "_vs_graphrag" in h:
            ws0.cell(row=summary_start, column=c).fill = gray

    baseline_headers = [
        "method",
        "runs",
        "mean_coverage",
        "mean_entropy_normalized",
        "mean_max_share",
        "mean_kl_selected_vs_graph",
        "mean_avg_query_similarity",
    ]
    baseline_rows: List[Dict[str, object]] = []
    for method in ("graphrag", "random"):
        m_rows = [r for r in rows if r.get("method") == method]
        if not m_rows:
            continue
        baseline_rows.append(
            {
                "method": method,
                "runs": len(m_rows),
                "mean_coverage": mean(float(r["coverage"]) for r in m_rows),
                "mean_entropy_normalized": mean(
                    float(r["entropy_normalized"]) for r in m_rows
                ),
                "mean_max_share": mean(float(r["max_share"]) for r in m_rows),
                "mean_kl_selected_vs_graph": mean(
                    float(r["kl_selected_vs_graph"]) for r in m_rows
                ),
                "mean_avg_query_similarity": mean(
                    float(r["avg_query_similarity"]) for r in m_rows
                ),
            }
        )
    ws0.cell(row=next_row + 1, column=1, value="Cumulative Baselines")
    _write_table(ws0, next_row + 2, baseline_headers, baseline_rows)

    all_headers = [
        "dataset_id",
        "dataset_seed",
        "query_id",
        "query_seed",
        "top_k",
        "seed_k",
        "lambda",
        "method",
        "communities_total",
        "selected_k",
        "represented_count",
        "coverage",
        "entropy",
        "entropy_normalized",
        "max_share",
        "kl_selected_vs_graph",
        "avg_query_similarity",
        "lift_coverage_vs_graphrag",
        "lift_coverage_vs_random",
        "lift_entropy_vs_graphrag",
        "lift_entropy_vs_random",
        "delta_entropy_vs_graphrag",
        "delta_entropy_vs_random",
        "improve_max_share_vs_graphrag",
        "improve_max_share_vs_random",
        "improve_kl_vs_graphrag",
        "improve_kl_vs_random",
        "sim_retention_vs_graphrag",
        "sim_retention_vs_random",
        "pass_diversity_vs_graphrag",
        "pass_diversity_vs_random",
        "pass_relevance_vs_graphrag",
        "pass_relevance_vs_random",
        "pass_overall_vs_graphrag",
        "pass_overall_vs_random",
    ]

    metric_check_cols = [
        "lift_coverage_vs_graphrag",
        "lift_coverage_vs_random",
        "lift_entropy_vs_graphrag",
        "lift_entropy_vs_random",
        "delta_entropy_vs_graphrag",
        "delta_entropy_vs_random",
        "improve_max_share_vs_graphrag",
        "improve_max_share_vs_random",
        "improve_kl_vs_graphrag",
        "improve_kl_vs_random",
        "sim_retention_vs_graphrag",
        "sim_retention_vs_random",
    ]
    bool_cols = [h for h in all_headers if h.startswith("pass_")]

    dataset_ids = sorted({str(r["dataset_id"]) for r in rows})
    for i, dataset_id in enumerate(dataset_ids, start=1):
        ws = wb.create_sheet(title=_sheet_title(f"{i:02d}_{dataset_id}"))
        ds_rows = [r for r in rows if str(r["dataset_id"]) == dataset_id]
        ds_rows.sort(
            key=lambda r: (
                str(r["query_id"]),
                0
                if r["method"] == "graphrag"
                else 1
                if r["method"] == "random"
                else 2,
                -1.0 if r["lambda"] == "NA" else float(r["lambda"]),
            )
        )
        _write_table(ws, 1, all_headers, ds_rows)
        for c, h in enumerate(all_headers, start=1):
            if "_vs_graphrag" in h:
                ws.cell(row=1, column=c).fill = gray

        header_index = {h: idx + 1 for idx, h in enumerate(all_headers)}
        for ridx, row in enumerate(ds_rows, start=2):
            metric_status = _metric_pass_statuses(row, thresholds)
            for col_name in metric_check_cols:
                cidx = header_index[col_name]
                status = metric_status[col_name]
                if status is None:
                    continue
                ws.cell(row=ridx, column=cidx).fill = green if status else red

            for col_name in bool_cols:
                cidx = header_index[col_name]
                status = _bool_status(row[col_name])
                if status is None:
                    continue
                ws.cell(row=ridx, column=cidx).fill = green if status else red

        _auto_width(ws)

    _auto_width(ws0)
    out = Path(args.output_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
