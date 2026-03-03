from __future__ import annotations

import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path
from statistics import mean
from typing import Dict, List, Tuple


def _load_rows(path: Path) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    with path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            parsed: Dict[str, object] = {}
            for k, v in row.items():
                if v in ("N/A", "", None):
                    parsed[k] = "N/A"
                    continue
                if k.startswith("pass_"):
                    parsed[k] = str(v).strip().lower() == "true"
                    continue
                if k in {
                    "dataset_seed",
                    "query_seed",
                    "top_k",
                    "seed_k",
                    "communities_total",
                    "selected_k",
                    "represented_count",
                }:
                    parsed[k] = int(float(v))
                    continue
                if k in {"dataset_id", "query_id", "method"}:
                    parsed[k] = v
                    continue
                if k == "lambda":
                    parsed[k] = v if v == "NA" else float(v)
                    continue
                parsed[k] = float(v)
            if (
                "delta_query_similarity_vs_graphrag" not in parsed
                and "sim_retention_vs_graphrag" in parsed
            ):
                parsed["delta_query_similarity_vs_graphrag"] = parsed[
                    "sim_retention_vs_graphrag"
                ]
            if (
                "delta_query_similarity_vs_random" not in parsed
                and "sim_retention_vs_random" in parsed
            ):
                parsed["delta_query_similarity_vs_random"] = parsed[
                    "sim_retention_vs_random"
                ]
            rows.append(parsed)
    return rows


def _to_float_or_none(value: object) -> float | None:
    if isinstance(value, (float, int)):
        return float(value)
    return None


def _auto_width(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))


def _mean_or_na(values: List[float]) -> float | str:
    if not values:
        return "N/A"
    return float(mean(values))


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Create a condensed cumulative-only Excel report for Bunny behavior metrics."
        )
    )
    parser.add_argument(
        "--rows-csv",
        default="synthetic_bunny/output/behavior_runner/behavior_results_rows.csv",
        help="Path to behavior_results_rows.csv.",
    )
    parser.add_argument(
        "--metadata-json",
        default="synthetic_bunny/output/behavior_runner/behavior_metadata.json",
        help="Path to behavior_metadata.json.",
    )
    parser.add_argument(
        "--output-xlsx",
        default="presentation/testing/behavior_report_condensed.xlsx",
        help="Output workbook path.",
    )
    args = parser.parse_args()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl. Install with: pip install openpyxl"
        ) from exc

    rows = _load_rows(Path(args.rows_csv))
    meta = json.loads(Path(args.metadata_json).read_text(encoding="utf-8"))
    thresholds: Dict[str, float] = meta.get("thresholds", {})
    min_sim_delta_gr = thresholds.get(
        "min_delta_query_similarity_vs_graphrag",
        thresholds.get("min_sim_retention_vs_graphrag", 0.0),
    )
    min_sim_delta_rd = thresholds.get(
        "min_delta_query_similarity_vs_random",
        thresholds.get("min_sim_retention_vs_random", 0.0),
    )
    lambdas = [float(x) for x in meta.get("lambdas", [])]
    params = meta.get("params", {})

    pass_fill = PatternFill(fill_type="solid", fgColor="FFBDD7EE")
    fail_fill = PatternFill(fill_type="solid", fgColor="FFFCE4D6")
    gray = PatternFill(fill_type="solid", fgColor="FFEFEFEF")

    # Baseline lookup by (dataset_id, query_id, method) for coverage deltas.
    baseline_coverage: Dict[Tuple[str, str, str], float] = {}
    for row in rows:
        method = str(row.get("method"))
        if method not in ("graphrag", "random"):
            continue
        key = (str(row["dataset_id"]), str(row["query_id"]), method)
        baseline_coverage[key] = float(row["coverage"])

    bunny_rows = [r for r in rows if r.get("method") == "bunny"]
    baseline_rows = [r for r in rows if r.get("method") in ("graphrag", "random")]
    by_lambda: Dict[float, List[Dict[str, object]]] = defaultdict(list)
    for row in bunny_rows:
        by_lambda[float(row["lambda"])].append(row)

    summary_headers = [
        "lambda",
        "runs",
        "mean_lift_entropy_vs_graphrag",
        "mean_lift_entropy_vs_random",
        "mean_delta_coverage_vs_graphrag",
        "mean_delta_coverage_vs_random",
        "mean_improve_kl_vs_graphrag",
        "mean_improve_kl_vs_random",
        "mean_delta_query_similarity_vs_graphrag",
        "mean_delta_query_similarity_vs_random",
    ]

    summary_rows: List[Dict[str, object]] = []
    for lam in sorted(lambdas):
        lam_rows = by_lambda.get(lam, [])
        if not lam_rows:
            continue

        lift_entropy_gr = [
            float(r["lift_entropy_vs_graphrag"])
            for r in lam_rows
            if isinstance(r["lift_entropy_vs_graphrag"], (float, int))
        ]
        lift_entropy_rd = [
            float(r["lift_entropy_vs_random"])
            for r in lam_rows
            if isinstance(r["lift_entropy_vs_random"], (float, int))
        ]

        delta_cov_gr: List[float] = []
        delta_cov_rd: List[float] = []
        for r in lam_rows:
            key_gr = (str(r["dataset_id"]), str(r["query_id"]), "graphrag")
            key_rd = (str(r["dataset_id"]), str(r["query_id"]), "random")
            b_cov = float(r["coverage"])
            if key_gr in baseline_coverage:
                delta_cov_gr.append(b_cov - baseline_coverage[key_gr])
            if key_rd in baseline_coverage:
                delta_cov_rd.append(b_cov - baseline_coverage[key_rd])

        summary_rows.append(
            {
                "lambda": lam,
                "runs": len(lam_rows),
                "mean_lift_entropy_vs_graphrag": _mean_or_na(lift_entropy_gr),
                "mean_lift_entropy_vs_random": _mean_or_na(lift_entropy_rd),
                "mean_delta_coverage_vs_graphrag": _mean_or_na(delta_cov_gr),
                "mean_delta_coverage_vs_random": _mean_or_na(delta_cov_rd),
                "mean_improve_kl_vs_graphrag": mean(
                    float(r["improve_kl_vs_graphrag"]) for r in lam_rows
                ),
                "mean_improve_kl_vs_random": mean(
                    float(r["improve_kl_vs_random"]) for r in lam_rows
                ),
                "mean_delta_query_similarity_vs_graphrag": mean(
                    float(r["delta_query_similarity_vs_graphrag"]) for r in lam_rows
                ),
                "mean_delta_query_similarity_vs_random": mean(
                    float(r["delta_query_similarity_vs_random"]) for r in lam_rows
                ),
            }
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "00_cumulative"

    ws["A1"] = "Run Setup"
    ws["A2"] = "n"
    ws["B2"] = params.get("n")
    ws["A3"] = "dim"
    ws["B3"] = params.get("dim")
    ws["A4"] = "scale_prob"
    ws["B4"] = params.get("scale_prob")

    ws["D1"] = "Legend"
    ws["D2"] = "Pass"
    ws["E2"] = "Blue"
    ws["D3"] = "Fail"
    ws["E3"] = "Orange"
    ws["E2"].fill = pass_fill
    ws["E3"].fill = fail_fill

    ws["H1"] = "Thresholds"
    threshold_rows = [
        ("min_lift_entropy_vs_graphrag", thresholds.get("min_lift_entropy_vs_graphrag")),
        ("min_lift_entropy_vs_random", thresholds.get("min_lift_entropy_vs_random")),
        ("min_improve_kl_vs_graphrag", thresholds.get("min_improve_kl_vs_graphrag")),
        ("min_improve_kl_vs_random", thresholds.get("min_improve_kl_vs_random")),
        ("min_delta_query_similarity_vs_graphrag", min_sim_delta_gr),
        ("min_delta_query_similarity_vs_random", min_sim_delta_rd),
        ("min_delta_coverage_vs_graphrag", 0.0),
        ("min_delta_coverage_vs_random", 0.0),
    ]
    for i, (k, v) in enumerate(threshold_rows, start=2):
        ws.cell(row=i, column=8, value=k)
        ws.cell(row=i, column=9, value=v)

    baseline_summary: List[Dict[str, object]] = []
    for method in ("graphrag", "random"):
        m_rows = [r for r in baseline_rows if r.get("method") == method]
        if not m_rows:
            continue
        baseline_summary.append(
            {
                "method": method,
                "runs": len(m_rows),
                "mean_coverage": mean(float(r["coverage"]) for r in m_rows),
                "mean_entropy_normalized": mean(
                    float(r["entropy_normalized"]) for r in m_rows
                ),
                "mean_kl_selected_vs_graph": mean(
                    float(r["kl_selected_vs_graph"]) for r in m_rows
                ),
                "mean_avg_query_similarity": mean(
                    float(r["avg_query_similarity"]) for r in m_rows
                ),
            }
        )

    ws["A6"] = "Baseline Values"
    baseline_headers = [
        "method",
        "runs",
        "mean_coverage",
        "mean_entropy_normalized",
        "mean_kl_selected_vs_graph",
        "mean_avg_query_similarity",
    ]
    for c, h in enumerate(baseline_headers, start=1):
        ws.cell(row=7, column=c, value=h)
    for ridx, row in enumerate(baseline_summary, start=8):
        for cidx, h in enumerate(baseline_headers, start=1):
            ws.cell(row=ridx, column=cidx, value=row[h])

    start = 12
    ws.cell(row=start - 1, column=1, value="Lambda Summary")
    for c, h in enumerate(summary_headers, start=1):
        ws.cell(row=start, column=c, value=h)
        if "_vs_graphrag" in h:
            ws.cell(row=start, column=c).fill = gray

    for ridx, row in enumerate(summary_rows, start=start + 1):
        for cidx, h in enumerate(summary_headers, start=1):
            ws.cell(row=ridx, column=cidx, value=row[h])

        checks = {
            "mean_lift_entropy_vs_graphrag": (
                None
                if _to_float_or_none(row["mean_lift_entropy_vs_graphrag"]) is None
                else float(row["mean_lift_entropy_vs_graphrag"])
                >= thresholds["min_lift_entropy_vs_graphrag"]
            ),
            "mean_lift_entropy_vs_random": (
                None
                if _to_float_or_none(row["mean_lift_entropy_vs_random"]) is None
                else float(row["mean_lift_entropy_vs_random"])
                >= thresholds["min_lift_entropy_vs_random"]
            ),
            "mean_delta_coverage_vs_graphrag": float(
                row["mean_delta_coverage_vs_graphrag"]
            )
            >= 0.0,
            "mean_delta_coverage_vs_random": float(
                row["mean_delta_coverage_vs_random"]
            )
            >= 0.0,
            "mean_improve_kl_vs_graphrag": float(row["mean_improve_kl_vs_graphrag"])
            >= thresholds["min_improve_kl_vs_graphrag"],
            "mean_improve_kl_vs_random": float(row["mean_improve_kl_vs_random"])
            >= thresholds["min_improve_kl_vs_random"],
            "mean_delta_query_similarity_vs_graphrag": float(
                row["mean_delta_query_similarity_vs_graphrag"]
            )
            >= min_sim_delta_gr,
            "mean_delta_query_similarity_vs_random": float(
                row["mean_delta_query_similarity_vs_random"]
            )
            >= min_sim_delta_rd,
        }
        header_idx = {h: i + 1 for i, h in enumerate(summary_headers)}
        for metric, status in checks.items():
            if status is None:
                continue
            ws.cell(row=ridx, column=header_idx[metric]).fill = (
                pass_fill if status else fail_fill
            )

    _auto_width(ws)
    out = Path(args.output_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
